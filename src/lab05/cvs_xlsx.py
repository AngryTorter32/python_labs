import csv
from pathlib import Path
from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    try:
        path_c = Path(csv_path)
        path_x = Path(xlsx_path)
    except FileNotFoundError:
        raise FileNotFoundError
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    with open(path_c, encoding='utf-8') as f: #по возможности добавить проверку на адекватность csv файла
        for row in csv.reader(f):
            ws.append(row)
    wb.save(path_x)


csv_to_xlsx('C:\\Users\\kuzne\\Desktop\\laby_piton\\python_labs\\data\\samples\\people.csv',
            'C:\\Users\\kuzne\\Desktop\\laby_piton\\python_labs\\data\\out\\people.xlsx')
