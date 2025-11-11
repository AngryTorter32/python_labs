import csv
from pathlib import Path
from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

'''
def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    try:
        path_c = Path(csv_path)
        path_x = Path(xlsx_path)
    except FileNotFoundError:
        raise FileNotFoundError
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    with open(path_c, encoding='utf-8') as f:
        for row in csv.reader(f):
            ws.append(row)
    wb.save(path_x)
'''

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None: #чет накодил
    try:
        path_c = Path(csv_path)
        path_x = Path(xlsx_path)
    except FileNotFoundError:
        raise FileNotFoundError
    with open(path_c, encoding='utf-8') as f:
        try:
            data = list(csv.DictReader(f))
        except csv.Error:
            raise ValueError
    if data == []:
        raise ValueError
    try:
        wb = load_workbook(path_x)
    except FileNotFoundError:
        raise FileNotFoundError
    ws = wb.active
    ws = data
    wb.save(path_x)

'''
def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    try:
        path_c = Path(csv_path)
        path_x = Path(xlsx_path)
    except FileNotFoundError:
        raise FileNotFoundError
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    with open(path_c, encoding='utf-8') as f:
        ws.append for row in csv.reader(f)
    wb.save(path_x)

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    try:
        path_c = Path(csv_path)
        path_x = Path(xlsx_path)
    except FileNotFoundError:
        raise FileNotFoundError
    with open(path_c, encoding='utf-8') as f:
        try:
            data = list(csv.DictReader(f))
        except csv.Error:
            raise ValueError
    if data == []:
        raise ValueError
    ws = data
    wb.save(path_x)
'''

csv_to_xlsx('C:\\Users\\kuzne\\Desktop\\laby_piton\\python_labs\\data\\samples\\people.csv',
            'C:\\Users\\kuzne\\Desktop\\laby_piton\\python_labs\\data\\out\\people.xlsx')